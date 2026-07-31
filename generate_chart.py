#!/usr/bin/env python3
"""从“基础数据.xlsx”生成 1–8 指标对比图。

布局对应手绘草图：
  上方（从上到下）：8 百KM模耗、7 换模率、4 异常单套平均KM、3 异常占比
  下方：1 退库总套数 vs 2 异常套数、5 使用总枚数 vs 6 换模数量
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import et_xmlfile  # noqa: F401 — openpyxl 的隐式依赖
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.ticker import MaxNLocator
from openpyxl.utils.datetime import from_excel


def _get_base_dir() -> Path:
    """获取 exe（frozen）或脚本所在目录。"""
    if getattr(sys, 'frozen', False):
        return Path(sys.argv[0]).parent
    return Path(__file__).parent


BASE_DIR = _get_base_dir()
DEFAULT_INPUT = BASE_DIR / "基础数据.xlsx"
DEFAULT_OUTPUT_DIR = BASE_DIR

METRICS = [
    "退库-总套数",
    "异常套数",
    "单套平均公里数",
    "异常退库占比",
    "模具使用总枚数",
    "换模数量",
    "换模率",
    "百KM模耗",
]

# 旧版数据文件中的指标名 → 当前脚本使用的指标名。
# 兼容 2026-07-30 之前生成的基础数据.xlsx（旧行名），避免 load_data 找不到指标。
NAME_ALIASES = {
    "异常单套平均KM": "单套平均公里数",
    "异常占比": "异常退库占比",
    "使用总枚数": "模具使用总枚数",
}


def choose_chinese_font() -> str:
    """选择本机可用中文字体，避免图中中文变成方框。"""
    preferred = [
        "PingFang SC",
        "Hiragino Sans GB",
        "Microsoft YaHei",
        "Noto Sans CJK SC",
        "SimHei",
        "Arial Unicode MS",
    ]
    installed = {f.name for f in fm.fontManager.ttflist}
    return next((name for name in preferred if name in installed), "DejaVu Sans")


def load_data(path: Path) -> tuple[list, dict[str, np.ndarray], str, dict[str, str]]:
    """读取首个工作表。

    列布局约定（2026-07-31 模板）：
      A 列: 规格 | B 列: 客户/总标题(B2=图片大标题) | C 列: 项目名称（图例/坐标轴显示以此列为准）
      D 列: 统计项目（数据行名以此列优先，用户可在此改指标名）
      E 列起: 日期数据（自动跳过文本表头）
    返回 (dates, rows, title, display_names)：
      rows 的 key 为数据行名映射后的规范名（METRICS），display_names 为其对应的 C 列项目名称。
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    # 图片大标题取 B2 单元格内容
    title = sheet.cell(2, 2).value
    if not title:
        title = "统计项目综合对比分析"
    title = str(title).strip()

    # 数据列从 D 列(4)开始扫描，但只接受数值型表头（Excel 日期序列号）。
    # 新模板中 D 列是“统计项目”文本，与 C 列重复，必须跳过；旧模板 D 列即首个日期。
    date_columns = [
        col
        for col in range(4, sheet.max_column + 1)
        if isinstance(sheet.cell(1, col).value, (int, float))
    ]
    if not date_columns:
        raise ValueError("第一行未找到日期列（D 列起需为数值型日期表头）。")
    dates = [sheet.cell(1, col).value for col in date_columns]
    dates = [
        from_excel(value, workbook.epoch) if isinstance(value, (int, float)) else value
        for value in dates
    ]
    rows: dict[str, np.ndarray] = {}
    display_names: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        # 数据行名：D 列（统计项目）优先，空则回退 C 列（项目名称）
        row_name = sheet.cell(row, 4).value or sheet.cell(row, 3).value
        # 图例/坐标轴显示名：C 列（项目名称）优先，空则回退 D 列
        display_name = sheet.cell(row, 3).value or sheet.cell(row, 4).value
        if row_name:
            # 旧版行名 → 当前指标名（兼容旧数据文件）
            canonical = NAME_ALIASES.get(str(row_name), str(row_name))
            values = [
                sheet.cell(row, col).value for col in date_columns
            ]
            rows[canonical] = np.asarray(
                [
                    np.nan if value in (None, "") else float(value)
                    for value in values
                ],
                dtype=float,
            )
            display_names[canonical] = str(display_name) if display_name else canonical

    missing = [name for name in METRICS if name not in rows]
    if missing:
        raise ValueError(f"工作簿缺少指标：{', '.join(missing)}")
    if not dates or any(value is None for value in dates):
        raise ValueError("日期表头为空或不完整。")
    if any(len(rows[name]) != len(dates) for name in METRICS):
        raise ValueError("指标数据长度与日期列数不一致。")

    # 百分比单元格有时会在源数据缺失时计算为 0；按其基础指标判定为空，
    # 避免把“无数据”误画成 0%，并让折线在该日期真正断开。
    rows["异常退库占比"][
        ~np.isfinite(rows["异常套数"])
    ] = np.nan
    rows["换模率"][
        ~np.isfinite(rows["换模数量"])
        | ~np.isfinite(rows["模具使用总枚数"])
    ] = np.nan

    return dates, rows, title, display_names


def add_line_panel(
    ax: plt.Axes,
    dates: list,
    values: np.ndarray,
    number: int,
    label: str,
    color: str,
    percentage: bool = False,
) -> None:
    """绘制一条与手绘草图相呼应的横向趋势线。"""
    ax.plot(
        dates,
        values,
        color=color,
        linewidth=2.4,
        marker="o",
        markersize=5,
        markerfacecolor="white",
        markeredgewidth=1.8,
        zorder=3,
    )
    ax.fill_between(dates, values, np.nanmin(values), color=color, alpha=0.07)
    ax.text(
        -0.055,
        0.50,
        str(number),
        transform=ax.transAxes,
        ha="center",
        va="center",
        fontsize=12,
        fontweight="bold",
        color="white",
        bbox=dict(boxstyle="circle,pad=0.32", facecolor=color, edgecolor="none"),
    )
    ax.text(
        0.005,
        0.86,
        label,
        transform=ax.transAxes,
        ha="left",
        va="top",
        fontsize=10.5,
        fontweight="bold",
        color="#263238",
    )

    value_min = float(np.nanmin(values))
    value_max = float(np.nanmax(values))
    span = value_max - value_min
    pad = max(span * 0.24, abs(float(np.nanmean(values))) * 0.05, 0.02)
    ax.set_ylim(value_min - pad, value_max + pad)
    ax.grid(axis="y", color="#DCE3EA", linewidth=0.8, alpha=0.8)
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)
    ax.tick_params(axis="x", length=0, labelbottom=False)
    ax.tick_params(axis="y", length=0, labelsize=8, colors="#66727D")
    if percentage:
        ax.yaxis.set_major_formatter(lambda x, _: f"{x:.0%}")
    else:
        ax.yaxis.set_major_formatter(lambda x, _: f"{x:,.2f}".rstrip("0").rstrip("."))

    finite_indices = np.flatnonzero(np.isfinite(values))
    if finite_indices.size == 0:
        return
    last_index = int(finite_indices[-1])
    last_value = values[last_index]
    last_text = f"{last_value:.0%}" if percentage else f"{last_value:,.2f}".rstrip("0").rstrip(".")
    ax.annotate(
        last_text,
        (dates[last_index], last_value),
        xytext=(7, 0),
        textcoords="offset points",
        va="center",
        fontsize=8.5,
        fontweight="bold",
        color=color,
    )


def add_stacked_comparison(
    ax: plt.Axes,
    dates: list,
    total: np.ndarray,
    subset: np.ndarray,
    number_pair: str,
    title: str,
    total_label: str,
    subset_label: str,
    base_color: str,
    subset_color: str,
) -> None:
    """用“子集 + 其余 = 总量”的叠加柱直接比较总量与重点子集。"""
    remainder = total - subset
    if np.any(remainder < 0):
        raise ValueError(f"{subset_label}存在大于{total_label}的值，无法画成子集叠加柱。")

    x = np.arange(len(dates))
    bars_subset = ax.bar(
        x,
        subset,
        width=0.68,
        color=subset_color,
        edgecolor="#334155",
        linewidth=0.6,
        label=subset_label,
        zorder=3,
    )
    ax.bar(
        x,
        remainder,
        width=0.68,
        bottom=subset,
        color=base_color,
        edgecolor="#334155",
        linewidth=0.6,
        label=f"{total_label}中的其余部分",
        zorder=2,
    )
    ax.text(
        0.0,
        1.10,
        number_pair,
        transform=ax.transAxes,
        fontsize=11,
        fontweight="bold",
        color="white",
        bbox=dict(boxstyle="round,pad=0.28", facecolor="#334155", edgecolor="none"),
    )
    ax.text(
        0.13,
        1.10,
        title,
        transform=ax.transAxes,
        fontsize=11,
        fontweight="bold",
        va="center",
        color="#263238",
    )

    for i, (bar, total_value, subset_value) in enumerate(
        zip(bars_subset, total, subset, strict=True)
    ):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            total_value + max(total) * 0.025,
            f"{total_value:,.0f}",
            ha="center",
            va="bottom",
            fontsize=7.5,
            color="#334155",
        )
        if subset_value > 0:
            ax.text(
                i,
                subset_value / 2,
                f"{subset_value:,.0f}",
                ha="center",
                va="center",
                fontsize=7.5,
                fontweight="bold",
                color="white",
            )

    ax.set_xticks(x, [d.strftime("%m-%d") for d in dates], rotation=40, ha="right")
    ax.set_ylim(0, max(total) * 1.19)
    ax.grid(axis="y", color="#DCE3EA", linewidth=0.8, alpha=0.8, zorder=0)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#B7C1CB")
    ax.tick_params(axis="y", length=0, labelsize=8, colors="#66727D")
    ax.tick_params(axis="x", length=0, labelsize=8, colors="#66727D")
    ax.yaxis.set_major_formatter(lambda x, _: f"{x:,.0f}")
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(0, 1.02),
        frameon=False,
        ncol=2,
        fontsize=8,
        handlelength=1.4,
        columnspacing=1.0,
    )


def add_combined_trends(ax: plt.Axes, dates: list, data: dict[str, np.ndarray], display_names: dict[str, str]) -> None:
    """把不同单位的四项趋势统一换算为首日=100后合并展示。"""
    specs = [
        (8, "百KM模耗", "#4062A6", ""),
        (7, "换模率", "#D28B26", "%"),
        (4, "单套平均公里数", "#657A4B", " KM"),
        (3, "异常退库占比", "#B85C72", "%"),
    ]
    for number, name, color, unit in specs:
        raw = data[name]
        indexed = raw / raw[0] * 100
        ax.plot(
            dates,
            indexed,
            linewidth=2.5,
            marker="o",
            markersize=5,
            markerfacecolor="white",
            markeredgewidth=1.7,
            color=color,
            label=f"{number} {display_names.get(name, name)}",
        )
        if unit == "%":
            latest = f"{raw[-1]:.0%}"
        elif unit == " KM":
            latest = f"{raw[-1]:,.0f} KM"
        else:
            latest = f"{raw[-1]:.2f}"
        y_offset = {3: 10, 4: -9}.get(number, 0)
        ax.annotate(
            f"{number}  {latest}",
            (dates[-1], indexed[-1]),
            xytext=(8, y_offset),
            textcoords="offset points",
            va="center",
            fontsize=8.5,
            fontweight="bold",
            color=color,
        )

    ax.axhline(100, color="#7B8794", linestyle="--", linewidth=1.0, alpha=0.8)
    ax.text(
        dates[0],
        102,
        "首日基准 = 100",
        fontsize=8,
        color="#66727D",
        ha="left",
        va="bottom",
    )
    ax.set_title(
        "趋势指标合并对比（3、4、7、8）",
        loc="left",
        fontsize=13,
        fontweight="bold",
        color="#263238",
        pad=14,
    )
    ax.set_ylabel("指数（首日=100）", fontsize=9, color="#5E6B76")
    ax.grid(axis="y", color="#DCE3EA", linewidth=0.8)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#B7C1CB")
    ax.tick_params(axis="both", length=0, labelsize=8.5, colors="#66727D")
    ax.set_xticks(dates, [d.strftime("%m-%d") for d in dates])
    ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.5, 1.01),
        frameon=False,
        ncol=4,
        fontsize=8.5,
    )
    ax.margins(x=0.05)


def add_combined_bars(ax: plt.Axes, dates: list, data: dict[str, np.ndarray], display_names: dict[str, str]) -> None:
    """在同一面板中并排展示两组“子集占总量”叠加柱。"""
    return_total = data["退库-总套数"]
    abnormal = data["异常套数"]
    use_total = data["模具使用总枚数"]
    changes = data["换模数量"]

    abnormal_share = abnormal / return_total
    change_share = changes / use_total
    x = np.arange(len(dates))
    width = 0.34
    return_x = x - width / 2
    use_x = x + width / 2

    ax.bar(
        return_x,
        abnormal_share,
        width,
        color="#C65359",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("异常套数", "异常套数"),
        zorder=3,
    )
    ax.bar(
        return_x,
        1 - abnormal_share,
        width,
        bottom=abnormal_share,
        color="#9FB7DA",
        edgecolor="#334155",
        linewidth=0.6,
        label="正常退库",
        zorder=2,
    )
    ax.bar(
        use_x,
        change_share,
        width,
        color="#D27B35",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("换模数量", "换模数量"),
        zorder=3,
    )
    ax.bar(
        use_x,
        1 - change_share,
        width,
        bottom=change_share,
        color="#A9C5B2",
        edgecolor="#334155",
        linewidth=0.6,
        label="未换模使用",
        zorder=2,
    )

    for pos, total, subset, share in zip(
        return_x, return_total, abnormal, abnormal_share, strict=True
    ):
        ax.text(
            pos,
            1.035,
            f"退库 {total:.0f}",
            ha="center",
            va="bottom",
            fontsize=7.2,
            color="#40505F",
            rotation=90,
        )
        ax.text(
            pos,
            max(share / 2, 0.035),
            f"{subset:.0f}\n{share:.0%}",
            ha="center",
            va="center",
            fontsize=6.8,
            fontweight="bold",
            color="white",
        )
    for pos, total, subset, share in zip(
        use_x, use_total, changes, change_share, strict=True
    ):
        ax.text(
            pos,
            1.035,
            f"使用 {total:.0f}",
            ha="center",
            va="bottom",
            fontsize=7.2,
            color="#40505F",
            rotation=90,
        )
        ax.text(
            pos,
            max(share / 2, 0.035),
            f"{subset:.0f}\n{share:.0%}",
            ha="center",
            va="center",
            fontsize=6.8,
            fontweight="bold",
            color="white",
        )

    ax.set_title(
        "总量与子集柱状图合并对比（1 vs 2、5 vs 6）",
        loc="left",
        fontsize=13,
        fontweight="bold",
        color="#263238",
        pad=14,
    )
    ax.set_ylabel("各自总量占比", fontsize=9, color="#5E6B76")
    ax.set_ylim(0, 1.25)
    ax.set_xticks(x, [d.strftime("%m-%d") for d in dates], rotation=35, ha="right")
    ax.yaxis.set_major_formatter(lambda value, _: f"{value:.0%}")
    ax.grid(axis="y", color="#DCE3EA", linewidth=0.8, zorder=0)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#B7C1CB")
    ax.tick_params(axis="both", length=0, labelsize=8.5, colors="#66727D")
    ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.5, 1.01),
        frameon=False,
        ncol=4,
        fontsize=8.5,
    )


def add_multi_axis_trends(
    ax_km: plt.Axes, dates: list, data: dict[str, np.ndarray], display_names: dict[str, str]
) -> None:
    """一张图中用三个纵轴呈现四条真实数值趋势。"""
    ax_rate = ax_km.twinx()
    ax_consumption = ax_km.twinx()
    ax_consumption.spines["right"].set_position(("axes", 1.09))
    ax_consumption.spines["right"].set_visible(True)

    line_km = ax_km.plot(
        dates,
        data["单套平均公里数"],
        color="#657A4B",
        linewidth=2.6,
        marker="o",
        markerfacecolor="white",
        markeredgewidth=1.8,
        label=display_names.get("单套平均公里数", "单套平均公里数"),
    )[0]
    line_abnormal = ax_rate.plot(
        dates,
        data["异常退库占比"],
        color="#B85C72",
        linewidth=2.4,
        marker="s",
        markerfacecolor="white",
        markeredgewidth=1.6,
        label=display_names.get("异常退库占比", "异常退库占比"),
    )[0]
    line_change = ax_rate.plot(
        dates,
        data["换模率"],
        color="#D28B26",
        linewidth=2.4,
        linestyle="--",
        marker="^",
        markerfacecolor="white",
        markeredgewidth=1.6,
        label=display_names.get("换模率", "换模率"),
    )[0]
    line_consumption = ax_consumption.plot(
        dates,
        data["百KM模耗"],
        color="#4062A6",
        linewidth=2.5,
        linestyle="-.",
        marker="D",
        markersize=5,
        markerfacecolor="white",
        markeredgewidth=1.6,
        label=display_names.get("百KM模耗", "百KM模耗"),
    )[0]

    ax_km.set_ylabel(display_names.get("单套平均公里数", "单套平均公里数"), color="#657A4B", fontsize=10)
    ax_rate.set_ylabel(display_names.get("异常退库占比", "异常退库占比") + " / " + display_names.get("换模率", "换模率"), color="#9B5364", fontsize=10)
    ax_consumption.set_ylabel(display_names.get("百KM模耗", "百KM模耗"), color="#4062A6", fontsize=10)
    ax_rate.yaxis.set_major_formatter(lambda value, _: f"{value:.0%}")
    ax_km.yaxis.set_major_formatter(lambda value, _: f"{value:,.0f}")
    ax_consumption.yaxis.set_major_formatter(
        lambda value, _: f"{value:.1f}".rstrip("0").rstrip(".")
    )

    ax_rate.set_ylim(0.05, 0.40)
    ax_consumption.set_ylim(1.0, 3.1)
    ax_km.set_ylim(500, 1650)
    ax_km.set_xticks(dates, [d.strftime("%m-%d") for d in dates], rotation=35, ha="right")
    ax_km.grid(axis="y", color="#DCE3EA", linewidth=0.8)
    ax_km.spines[["top", "right"]].set_visible(False)
    ax_rate.spines[["top", "left"]].set_visible(False)
    ax_consumption.spines[["top", "left"]].set_visible(False)
    ax_km.spines["left"].set_color("#657A4B")
    ax_rate.spines["right"].set_color("#B85C72")
    ax_consumption.spines["right"].set_color("#4062A6")
    ax_km.tick_params(axis="y", colors="#657A4B", length=0)
    ax_rate.tick_params(axis="y", colors="#9B5364", length=0)
    ax_consumption.tick_params(axis="y", colors="#4062A6", length=0)
    ax_km.tick_params(axis="x", colors="#66727D", length=0)

    ax_km.legend(
        [line_abnormal, line_km, line_change, line_consumption],
        [line.get_label() for line in [line_abnormal, line_km, line_change, line_consumption]],
        loc="upper center",
        bbox_to_anchor=(0.5, 1.11),
        frameon=False,
        ncol=4,
        fontsize=9,
    )


def add_dual_axis_grouped_bars(
    ax_return: plt.Axes, dates: list, data: dict[str, np.ndarray], display_names: dict[str, str]
) -> None:
    """一张双轴图中，每个日期并排展示两组叠加柱。"""
    ax_use = ax_return.twinx()
    x = np.arange(len(dates))
    width = 0.36
    return_x = x - width / 2
    use_x = x + width / 2

    return_total = data["退库-总套数"]
    abnormal = data["异常套数"]
    use_total = data["模具使用总枚数"]
    changes = data["换模数量"]

    return_bars = ax_return.bar(
        return_x,
        return_total,
        width,
        color="#9FB7DA",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("退库-总套数", "退库-总套数"),
        zorder=2,
    )
    abnormal_bars = ax_return.bar(
        return_x,
        abnormal,
        width * 0.58,
        color="#C65359",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("异常套数", "异常套数"),
        zorder=3,
    )
    use_bars = ax_use.bar(
        use_x,
        use_total,
        width,
        color="#A9C5B2",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("模具使用总枚数", "模具使用总枚数"),
        zorder=2,
    )
    change_bars = ax_use.bar(
        use_x,
        changes,
        width * 0.58,
        color="#D27B35",
        edgecolor="#334155",
        linewidth=0.6,
        label=display_names.get("换模数量", "换模数量"),
        zorder=3,
    )

    for position, total, subset in zip(return_x, return_total, abnormal, strict=True):
        if not np.isfinite(total):
            continue
        ax_return.text(
            position,
            total + 0.8,
            f"{total:.0f}",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#40505F",
        )
        if np.isfinite(subset):
            ax_return.text(
                position,
                subset / 2,
                f"{subset:.0f}",
                ha="center",
                va="center",
                fontsize=7.5,
                fontweight="bold",
                color="white",
            )
    for position, total, subset in zip(use_x, use_total, changes, strict=True):
        if not np.isfinite(total):
            continue
        ax_use.text(
            position,
            total + 5,
            f"{total:.0f}",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#40505F",
        )
        if np.isfinite(subset):
            ax_use.text(
                position,
                subset / 2,
                f"{subset:.0f}",
                ha="center",
                va="center",
                fontsize=7.5,
                fontweight="bold",
                color="white",
            )

    # 左右轴分别根据各自数据动态扩展，并为柱顶数据标签预留空间。
    return_axis_max = float(np.nanmax(np.concatenate([return_total, abnormal])))
    use_axis_max = float(np.nanmax(np.concatenate([use_total, changes])))
    ax_return.set_ylim(0, max(return_axis_max * 1.16, 1))
    ax_use.set_ylim(0, max(use_axis_max * 1.16, 1))
    ax_return.yaxis.set_major_locator(MaxNLocator(nbins=7, integer=True))
    ax_use.yaxis.set_major_locator(MaxNLocator(nbins=7, integer=True))
    ax_return.set_ylabel(display_names.get("退库-总套数", "退库-总套数") + " / " + display_names.get("异常套数", "异常套数"), color="#5D79A2", fontsize=10)
    ax_use.set_ylabel(display_names.get("模具使用总枚数", "模具使用总枚数") + " / " + display_names.get("换模数量", "换模数量"), color="#678B72", fontsize=10)
    ax_return.set_xticks(x, [d.strftime("%m-%d") for d in dates], rotation=35, ha="right")
    ax_return.grid(axis="y", color="#DCE3EA", linewidth=0.8, zorder=0)
    ax_return.spines[["top", "right"]].set_visible(False)
    ax_use.spines[["top", "left"]].set_visible(False)
    ax_return.spines["left"].set_color("#5D79A2")
    ax_use.spines["right"].set_color("#678B72")
    ax_return.tick_params(axis="y", colors="#5D79A2", length=0)
    ax_use.tick_params(axis="y", colors="#678B72", length=0)
    ax_return.tick_params(axis="x", colors="#66727D", length=0)
    ax_return.legend(
        [return_bars, abnormal_bars, use_bars, change_bars],
        [
            display_names.get("退库-总套数", "退库-总套数"),
            display_names.get("异常套数", "异常套数"),
            display_names.get("模具使用总枚数", "模具使用总枚数"),
            display_names.get("换模数量", "换模数量"),
        ],
        loc="upper center",
        bbox_to_anchor=(0.5, 1.11),
        frameon=False,
        ncol=4,
        fontsize=8.5,
    )


def add_percentage_trends(
    ax: plt.Axes, dates: list, data: dict[str, np.ndarray], display_names: dict[str, str]
) -> None:
    """在同一百分比坐标轴中展示异常退库占比与换模率。"""
    specs = [
        ("异常退库占比", "#B85C72", "s", "-"),
        ("换模率", "#D28B26", "^", "--"),
    ]
    all_values = np.concatenate([data[name] for name, *_ in specs])
    lower = max(0.0, float(np.nanmin(all_values)) - 0.04)
    upper = min(1.0, float(np.nanmax(all_values)) + 0.05)

    for series_index, (name, color, marker, line_style) in enumerate(specs):
        values = data[name]
        ax.plot(
            dates,
            values,
            color=color,
            linewidth=2.6,
            linestyle=line_style,
            marker=marker,
            markersize=6,
            markerfacecolor="white",
            markeredgewidth=1.7,
            label=display_names.get(name, name),
            zorder=3,
        )
        for point_index, (date, value) in enumerate(
            zip(dates, values, strict=True)
        ):
            if not np.isfinite(value):
                continue
            direction = 1 if (point_index + series_index) % 2 == 0 else -1
            ax.annotate(
                f"{value:.0%}",
                (date, value),
                xytext=(0, 8 * direction),
                textcoords="offset points",
                ha="center",
                va="bottom" if direction > 0 else "top",
                fontsize=8,
                color=color,
            )

    ax.set_ylim(lower, upper)
    ax.set_ylabel("比例 / 率", fontsize=10, color="#6B5860")
    ax.yaxis.set_major_formatter(lambda value, _: f"{value:.0%}")
    ax.set_xticks(
        dates,
        [date.strftime("%m-%d") for date in dates],
        rotation=35,
        ha="right",
    )
    ax.grid(axis="y", color="#DCE3EA", linewidth=0.8)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#B7C1CB")
    ax.tick_params(axis="both", length=0, colors="#66727D")
    ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.5, 1.10),
        frameon=False,
        ncol=2,
        fontsize=9,
    )


def build_chart(
    input_path: Path, output_dir: Path
) -> tuple[Path, Path, Path, Path, Path, Path]:
    dates, data, title, display_names = load_data(input_path)

    plt.rcParams.update(
        {
            "font.family": choose_chinese_font(),
            "axes.unicode_minus": False,
            "figure.facecolor": "#F7F9FC",
            "axes.facecolor": "#F7F9FC",
        }
    )
    output_path = output_dir / "统计项目综合对比分析图.png"
    svg_path = output_dir / "统计项目综合对比分析图.svg"
    fig = plt.figure(figsize=(15, 12), dpi=170)
    grid = fig.add_gridspec(
        2,
        1,
        height_ratios=[1.0, 1.05],
        hspace=0.34,
        left=0.08,
        right=0.88,
        top=0.89,
        bottom=0.06,
    )
    trend_axis = fig.add_subplot(grid[0, 0])
    add_multi_axis_trends(trend_axis, dates, data, display_names)

    bar_axis = fig.add_subplot(grid[1, 0])
    add_dual_axis_grouped_bars(bar_axis, dates, data, display_names)

    fig.suptitle(
        title,
        x=0.08,
        y=0.965,
        ha="left",
        fontsize=20,
        fontweight="bold",
        color="#18212B",
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, bbox_inches="tight", facecolor=fig.get_facecolor())
    fig.savefig(svg_path, bbox_inches="tight", facecolor=fig.get_facecolor(), format="svg")
    plt.close(fig)

    # 新增分面版：上方折线拆为 2 行 2 列，下方柱状图保持不变。
    faceted_png_path = output_dir / "统计项目综合对比分析图_折线分面版.png"
    faceted_svg_path = output_dir / "统计项目综合对比分析图_折线分面版.svg"
    faceted_fig = plt.figure(figsize=(15, 13), dpi=170)
    faceted_grid = faceted_fig.add_gridspec(
        3,
        2,
        height_ratios=[0.72, 0.72, 1.25],
        hspace=0.30,
        wspace=0.18,
        left=0.08,
        right=0.92,
        top=0.91,
        bottom=0.07,
    )
    line_specs = [
        (8, "百KM模耗", "#4062A6", False),
        (7, "换模率", "#D28B26", True),
        (4, "单套平均公里数", "#657A4B", False),
        (3, "异常退库占比", "#B85C72", True),
    ]
    for index, (number, name, color, percentage) in enumerate(line_specs):
        row, col = divmod(index, 2)
        axis = faceted_fig.add_subplot(faceted_grid[row, col])
        add_line_panel(
            axis,
            dates,
            data[name],
            number,
            display_names.get(name, name),
            color,
            percentage=percentage,
        )
        # 分面版不显示 1–8 编号，指标名称作为子图标题置于图框上方（用 C 列项目名称）。
        # 注意：add_line_panel 内部会在左上角画 label 文本，其内容等于 C 列显示名，
        # 必须一并移除，否则子图标题 + 左上角标签 = 同一名字出现两次。
        for artist in list(axis.texts):
            if artist.get_text() in {str(number), name, display_names.get(name, name)}:
                artist.remove()
        axis.set_title(
            display_names.get(name, name),
            loc="left",
            fontsize=11,
            fontweight="bold",
            color="#263238",
            pad=10,
        )
        axis.tick_params(axis="x", labelbottom=True)
        axis.set_xticks(
            dates,
            [date.strftime("%m-%d") for date in dates],
            rotation=35,
            ha="right",
        )
        # 最后一个点已由 add_line_panel 标注，这里补齐其余数据点标签。
        values = data[name]
        finite_indices = np.flatnonzero(np.isfinite(values))
        last_finite_index = int(finite_indices[-1]) if finite_indices.size else -1
        for point_index, (date, value) in enumerate(
            zip(dates, values, strict=True)
        ):
            if not np.isfinite(value) or point_index == last_finite_index:
                continue
            label = (
                f"{value:.0%}"
                if percentage
                else f"{value:,.2f}".rstrip("0").rstrip(".")
            )
            axis.annotate(
                label,
                (date, value),
                xytext=(0, 8 if point_index % 2 == 0 else -13),
                textcoords="offset points",
                ha="center",
                va="bottom" if point_index % 2 == 0 else "top",
                fontsize=7.2,
                color=color,
            )

    faceted_bar_axis = faceted_fig.add_subplot(faceted_grid[2, :])
    add_dual_axis_grouped_bars(faceted_bar_axis, dates, data, display_names)
    faceted_fig.suptitle(
        title,
        x=0.08,
        y=0.975,
        ha="left",
        fontsize=20,
        fontweight="bold",
        color="#18212B",
    )
    faceted_fig.savefig(
        faceted_png_path,
        bbox_inches="tight",
        facecolor=faceted_fig.get_facecolor(),
    )
    faceted_fig.savefig(
        faceted_svg_path,
        bbox_inches="tight",
        facecolor=faceted_fig.get_facecolor(),
        format="svg",
    )
    plt.close(faceted_fig)

    # 新增百分比双折线版：仅保留异常退库占比和换模率，共用一个纵轴。
    percentage_png_path = output_dir / "统计项目综合对比分析图_百分比双折线版.png"
    percentage_svg_path = output_dir / "统计项目综合对比分析图_百分比双折线版.svg"
    percentage_fig = plt.figure(figsize=(15, 10), dpi=170)
    percentage_grid = percentage_fig.add_gridspec(
        2,
        1,
        height_ratios=[0.8, 1.25],
        hspace=0.34,
        left=0.08,
        right=0.92,
        top=0.88,
        bottom=0.08,
    )
    percentage_axis = percentage_fig.add_subplot(percentage_grid[0, 0])
    add_percentage_trends(percentage_axis, dates, data, display_names)
    percentage_bar_axis = percentage_fig.add_subplot(percentage_grid[1, 0])
    add_dual_axis_grouped_bars(percentage_bar_axis, dates, data, display_names)
    percentage_fig.suptitle(
        title,
        x=0.08,
        y=0.965,
        ha="left",
        fontsize=20,
        fontweight="bold",
        color="#18212B",
    )
    percentage_fig.savefig(
        percentage_png_path,
        bbox_inches="tight",
        facecolor=percentage_fig.get_facecolor(),
    )
    percentage_fig.savefig(
        percentage_svg_path,
        bbox_inches="tight",
        facecolor=percentage_fig.get_facecolor(),
        format="svg",
    )
    plt.close(percentage_fig)
    return (
        output_path,
        svg_path,
        faceted_png_path,
        faceted_svg_path,
        percentage_png_path,
        percentage_svg_path,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成 1–8 指标对比分析图")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="输入 Excel 文件")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="两张 PNG 图片的输出目录",
    )
    return parser.parse_args()


if __name__ == "__main__":
    try:
        args = parse_args()
        build_chart(args.input, args.output_dir)
        print("✅ 图表生成完成！")
        print(f"   输出目录: {DEFAULT_OUTPUT_DIR}")
        input("按回车键退出...")
    except SystemExit:
        raise
    except Exception:
        import traceback

        error_log = BASE_DIR / "generate_chart_错误日志.txt"
        try:
            with open(error_log, "w", encoding="utf-8") as f:
                traceback.print_exc(file=f)
            print(f"❌ 程序出错，详情已写入: {error_log}")
        except Exception:
            traceback.print_exc()
        print("\n===== 错误信息 =====")
        traceback.print_exc()
        print("====================")
        input("程序异常退出，按回车键关闭窗口...")
